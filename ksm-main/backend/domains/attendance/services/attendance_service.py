#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Attendance Service untuk KSM Main Backend
Business logic untuk sistem absensi karyawan
"""

import logging
from datetime import datetime, date, time, timedelta
from typing import List, Dict, Optional, Tuple
import base64
import io
import json
from io import BytesIO
import pymysql
from sqlalchemy.exc import OperationalError

from config.database import db
from domains.attendance.models.attendance_models import (
    AttendanceRecord, AttendanceLeave, OvertimeRequest, AttendanceSettings
)
from domains.auth.models.auth_models import User
from domains.role.models.role_models import Department, UserRole, Role

# Import untuk export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    logging.warning("openpyxl not installed, Excel export will not work")

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
except ImportError:
    logging.warning("reportlab not installed, PDF export will not work")

logger = logging.getLogger(__name__)

def is_encryption_error(error: Exception) -> bool:
    """Cek apakah error terkait database encryption"""
    error_msg = str(error).lower()
    return (
        'encrypted' in error_msg or 
        'decryption' in error_msg or 
        'encryption management plugin' in error_msg or
        'error 192' in error_msg
    )

class AttendanceService:
    """Service untuk mengelola absensi karyawan"""
    
    def __init__(self):
        self.settings = None
    
    def get_attendance_settings(self) -> AttendanceSettings:
        """Ambil atau buat default attendance settings"""
        if self.settings is None:
            settings = AttendanceSettings.query.first()
            if not settings:
                settings = AttendanceSettings()
                db.session.add(settings)
                db.session.commit()
            self.settings = settings
        return self.settings
    
    def clock_in(self, user_id: int, latitude: float, longitude: float, 
                 address: str, photo_base64: str, notes: str = None) -> Dict:
        """Handle clock in dengan validasi"""
        try:
            # Cek apakah sudah clock in hari ini
            today = date.today()
            existing_record = AttendanceRecord.query.filter(
                AttendanceRecord.user_id == user_id,
                db.func.date(AttendanceRecord.clock_in) == today
            ).first()
            
            if existing_record:
                return {
                    'success': False,
                    'message': 'Anda sudah melakukan clock in hari ini',
                    'data': existing_record.to_dict()
                }
            
            # Ambil settings terlebih dahulu
            settings = self.get_attendance_settings()
            
            # Validasi foto jika diperlukan
            if settings.photo_required and not photo_base64:
                return {
                    'success': False,
                    'message': 'Foto wajib diambil saat clock in'
                }
            
            # Validasi ukuran foto
            if photo_base64:
                photo_size_mb = len(photo_base64.encode('utf-8')) / (1024 * 1024)
                if photo_size_mb > settings.max_photo_size_mb:
                    return {
                        'success': False,
                        'message': f'Ukuran foto terlalu besar. Maksimal {settings.max_photo_size_mb}MB'
                    }
            
            # Buat record baru
            now = datetime.now()
            record = AttendanceRecord(
                user_id=user_id,
                clock_in=now,
                clock_in_latitude=latitude,
                clock_in_longitude=longitude,
                clock_in_address=address,
                clock_in_photo=photo_base64,
                notes=notes,
                status='present'
            )
            
            # Tentukan status berdasarkan waktu dengan grace period
            if record.is_late(settings.work_start_time, settings.grace_period_minutes):
                record.status = 'late'
            
            db.session.add(record)
            db.session.commit()
            
            logger.info(f"User {user_id} clock in berhasil pada {now}")
            
            return {
                'success': True,
                'message': 'Clock in berhasil',
                'data': record.to_dict()
            }
            
        except Exception as e:
            logger.error(f"Error clock in: {str(e)}")
            db.session.rollback()
            return {
                'success': False,
                'message': f'Terjadi kesalahan: {str(e)}'
            }
    
    def clock_out(self, user_id: int, latitude: float, longitude: float, 
                  address: str, photo_base64: str, notes: str = None) -> Dict:
        """Handle clock out dengan kalkulasi durasi kerja"""
        try:
            logger.info(f"Clock out service - User ID: {user_id}")
            logger.info(f"Clock out service - Latitude: {latitude}, Longitude: {longitude}")
            logger.info(f"Clock out service - Address: {address}")
            logger.info(f"Clock out service - Photo present: {bool(photo_base64)}")
            logger.info(f"Clock out service - Notes: {notes}")
            
            # Cari record clock in hari ini
            today = date.today()
            record = AttendanceRecord.query.filter(
                AttendanceRecord.user_id == user_id,
                db.func.date(AttendanceRecord.clock_in) == today,
                AttendanceRecord.clock_out.is_(None)
            ).first()
            
            if not record:
                logger.warning(f"No clock in record found for user {user_id} on {today}")
                return {
                    'success': False,
                    'message': 'Anda belum melakukan clock in hari ini'
                }
            
            # Validasi foto jika diperlukan
            settings = self.get_attendance_settings()
            if settings.photo_required and not photo_base64:
                return {
                    'success': False,
                    'message': 'Foto wajib diambil saat clock out'
                }
            
            # Validasi task sebelum clock-out
            try:
                from domains.task.services.daily_task_service import DailyTaskService
                task_service = DailyTaskService()
                task_validation = task_service.validate_clockout_tasks(user_id, today)
                
                logger.info(f"Task validation result: {task_validation}")
                
                if not task_validation['valid']:
                    logger.warning(f"Task validation failed: {task_validation['message']}")
                    return {
                        'success': False,
                        'message': task_validation['message'],
                        'task_validation': task_validation,
                        'error_code': 'TASK_VALIDATION_FAILED'
                    }
            except ImportError:
                # Jika daily task service belum tersedia, skip validasi
                logger.warning("Daily task service not available, skipping task validation")
            except Exception as e:
                logger.error(f"Error validating tasks: {e}")
                # Continue with clock-out even if task validation fails
            
            # Update record dengan clock out
            now = datetime.now()
            record.clock_out = now
            record.clock_out_latitude = latitude
            record.clock_out_longitude = longitude
            record.clock_out_address = address
            record.clock_out_photo = photo_base64
            
            if notes:
                record.notes = f"{record.notes}\n{notes}" if record.notes else notes
            
            # Hitung durasi kerja
            record.calculate_work_duration()
            
            # Update status berdasarkan durasi
            record.status = record.get_status()
            
            db.session.commit()
            
            # Update attendance_id di daily tasks
            try:
                from domains.task.models.task_models import DailyTask
                DailyTask.query.filter(
                    DailyTask.user_id == user_id,
                    DailyTask.task_date == today,
                    DailyTask.attendance_id.is_(None)
                ).update({DailyTask.attendance_id: record.id})
                db.session.commit()
            except Exception as e:
                logger.error(f"Error updating task attendance_id: {e}")
                # Don't fail clock-out if task update fails
            
            logger.info(f"User {user_id} clock out berhasil pada {now}")
            
            return {
                'success': True,
                'message': 'Clock out berhasil',
                'data': record.to_dict()
            }
            
        except Exception as e:
            logger.error(f"Error clock out: {str(e)}")
            db.session.rollback()
            return {
                'success': False,
                'message': f'Terjadi kesalahan: {str(e)}'
            }
    
    def get_user_attendance(self, user_id: int, start_date: date = None, 
                           end_date: date = None, limit: int = 30) -> List[Dict]:
        """Ambil history absensi user"""
        try:
            query = AttendanceRecord.query.filter(AttendanceRecord.user_id == user_id)
            
            if start_date:
                query = query.filter(db.func.date(AttendanceRecord.clock_in) >= start_date)
            if end_date:
                query = query.filter(db.func.date(AttendanceRecord.clock_in) <= end_date)
            
            records = query.order_by(AttendanceRecord.clock_in.desc()).limit(limit).all()
            
            return [record.to_dict() for record in records]
            
        except Exception as e:
            logger.error(f"Error get user attendance: {str(e)}")
            return []
    
    def get_team_attendance(self, manager_user_id: int, start_date: date = None, 
                           end_date: date = None) -> List[Dict]:
        """Manager lihat absensi tim berdasarkan department"""
        try:
            # Cari department manager
            user_role = UserRole.query.filter(
                UserRole.user_id == manager_user_id,
                UserRole.is_active == True
            ).join(Role).filter(Role.is_management == True).first()
            
            if not user_role:
                return []
            
            # Cari semua user di department yang sama
            team_users = UserRole.query.filter(
                UserRole.role_id == user_role.role_id,
                UserRole.is_active == True
            ).all()
            
            user_ids = [ur.user_id for ur in team_users]
            
            # Ambil attendance records
            query = AttendanceRecord.query.filter(AttendanceRecord.user_id.in_(user_ids))
            
            if start_date:
                query = query.filter(db.func.date(AttendanceRecord.clock_in) >= start_date)
            if end_date:
                query = query.filter(db.func.date(AttendanceRecord.clock_in) <= end_date)
            
            records = query.order_by(AttendanceRecord.clock_in.desc()).all()
            
            return [record.to_dict() for record in records]
            
        except Exception as e:
            logger.error(f"Error get team attendance: {str(e)}")
            return []
    
    def get_all_attendance(self, start_date: date = None, end_date: date = None, 
                          user_id: int = None, status: str = None, search: str = None) -> List[Dict]:
        """Admin lihat semua absensi dengan filter"""
        try:
            query = AttendanceRecord.query
            
            if start_date:
                query = query.filter(db.func.date(AttendanceRecord.clock_in) >= start_date)
            if end_date:
                query = query.filter(db.func.date(AttendanceRecord.clock_in) <= end_date)
            if user_id:
                query = query.filter(AttendanceRecord.user_id == user_id)
            if status:
                query = query.filter(AttendanceRecord.status == status)
            
            # Tambahkan filter pencarian berdasarkan nama user dan alamat
            if search:
                # Join dengan tabel User untuk pencarian berdasarkan nama
                query = query.join(User, AttendanceRecord.user_id == User.id)
                search_term = f"%{search}%"
                query = query.filter(
                    db.or_(
                        User.username.ilike(search_term),
                        AttendanceRecord.clock_in_address.ilike(search_term),
                        AttendanceRecord.clock_out_address.ilike(search_term)
                    )
                )
            
            records = query.order_by(AttendanceRecord.clock_in.desc()).all()
            
            return [record.to_dict() for record in records]
            
        except Exception as e:
            logger.error(f"Error get all attendance: {str(e)}")
            return []
    
    def get_today_status(self, user_id: int = None) -> Dict:
        """Status absensi hari ini"""
        try:
            today = date.today()
            
            if user_id:
                # Status untuk user tertentu
                try:
                    record = AttendanceRecord.query.filter(
                        AttendanceRecord.user_id == user_id,
                        db.func.date(AttendanceRecord.clock_in) == today
                    ).first()
                    
                    if record:
                        return {
                            'user_id': user_id,
                            'has_clocked_in': True,
                            'has_clocked_out': record.clock_out is not None,
                            'clock_in_time': record.clock_in.isoformat() if record.clock_in else None,
                            'clock_out_time': record.clock_out.isoformat() if record.clock_out else None,
                            'status': record.status or 'present',
                            'work_duration_hours': round(record.work_duration_minutes / 60, 2) if record.work_duration_minutes else 0
                        }
                    else:
                        return {
                            'user_id': user_id,
                            'has_clocked_in': False,
                            'has_clocked_out': False,
                            'clock_in_time': None,
                            'clock_out_time': None,
                            'status': 'absent',
                            'work_duration_hours': 0
                        }
                except Exception as e:
                    logger.error(f"Error querying today status for user {user_id}: {str(e)}", exc_info=True)
                    return {
                        'user_id': user_id,
                        'has_clocked_in': False,
                        'has_clocked_out': False,
                        'clock_in_time': None,
                        'clock_out_time': None,
                        'status': 'absent',
                        'work_duration_hours': 0
                    }
            else:
                # Status untuk semua user (admin view)
                try:
                    total_users = User.query.filter(User.is_active == True).count()
                    clocked_in_users = AttendanceRecord.query.filter(
                        db.func.date(AttendanceRecord.clock_in) == today
                    ).count()
                    clocked_out_users = AttendanceRecord.query.filter(
                        db.func.date(AttendanceRecord.clock_in) == today,
                        AttendanceRecord.clock_out.isnot(None)
                    ).count()
                    
                    return {
                        'total_users': total_users,
                        'clocked_in_users': clocked_in_users,
                        'clocked_out_users': clocked_out_users,
                        'absent_users': total_users - clocked_in_users,
                        'attendance_rate': round((clocked_in_users / total_users * 100), 2) if total_users > 0 else 0
                    }
                except Exception as e:
                    logger.error(f"Error querying today status for all users: {str(e)}", exc_info=True)
                    return {
                        'total_users': 0,
                        'clocked_in_users': 0,
                        'clocked_out_users': 0,
                        'absent_users': 0,
                        'attendance_rate': 0
                    }
                
        except (OperationalError, pymysql.err.OperationalError) as e:
            if is_encryption_error(e):
                logger.error(f"Error get today status - Database encryption error: {str(e)}", exc_info=True)
                raise ValueError("DATABASE_ENCRYPTION_ERROR")
            logger.error(f"Error get today status - Database error: {str(e)}", exc_info=True)
            return {}
        except Exception as e:
            logger.error(f"Error get today status: {str(e)}", exc_info=True)
            return {}
    
    def submit_leave_request(self, user_id: int, leave_type: str, start_date: date, 
                           end_date: date, reason: str, medical_certificate: str = None,
                           emergency_contact: str = None, emergency_phone: str = None) -> Dict:
        """Pengajuan izin/sakit"""
        try:
            # Validasi tanggal
            if start_date > end_date:
                return {
                    'success': False,
                    'message': 'Tanggal mulai tidak boleh lebih besar dari tanggal selesai'
                }
            
            if start_date < date.today():
                return {
                    'success': False,
                    'message': 'Tanggal mulai tidak boleh di masa lalu'
                }
            
            # Hitung total hari
            total_days = (end_date - start_date).days + 1
            
            # Buat leave request
            leave_request = AttendanceLeave(
                user_id=user_id,
                leave_type=leave_type,
                start_date=start_date,
                end_date=end_date,
                total_days=total_days,
                reason=reason,
                medical_certificate=medical_certificate,
                emergency_contact=emergency_contact,
                emergency_phone=emergency_phone,
                status='pending'
            )
            
            db.session.add(leave_request)
            db.session.commit()
            
            logger.info(f"Leave request submitted by user {user_id}")
            
            return {
                'success': True,
                'message': 'Pengajuan izin berhasil dikirim',
                'data': leave_request.to_dict()
            }
            
        except Exception as e:
            logger.error(f"Error submit leave request: {str(e)}")
            db.session.rollback()
            return {
                'success': False,
                'message': f'Terjadi kesalahan: {str(e)}'
            }
    
    def approve_leave_request(self, leave_id: int, approver_id: int, 
                            action: str, rejection_reason: str = None) -> Dict:
        """Approval izin (approve/reject)"""
        try:
            leave_request = AttendanceLeave.query.get(leave_id)
            if not leave_request:
                return {
                    'success': False,
                    'message': 'Pengajuan izin tidak ditemukan'
                }
            
            if action == 'approve':
                leave_request.approve(approver_id)
                message = 'Pengajuan izin berhasil disetujui'
            elif action == 'reject':
                if not rejection_reason:
                    return {
                        'success': False,
                        'message': 'Alasan penolakan harus diisi'
                    }
                leave_request.reject(approver_id, rejection_reason)
                message = 'Pengajuan izin ditolak'
            else:
                return {
                    'success': False,
                    'message': 'Action tidak valid'
                }
            
            logger.info(f"Leave request {leave_id} {action}ed by {approver_id}")
            
            return {
                'success': True,
                'message': message,
                'data': leave_request.to_dict()
            }
            
        except Exception as e:
            logger.error(f"Error approve leave request: {str(e)}")
            db.session.rollback()
            return {
                'success': False,
                'message': f'Terjadi kesalahan: {str(e)}'
            }
    
    def submit_overtime_request(self, user_id: int, attendance_id: int, 
                              requested_hours: float, reason: str, 
                              task_description: str = None) -> Dict:
        """Pengajuan overtime"""
        try:
            # Validasi attendance record
            attendance = AttendanceRecord.query.get(attendance_id)
            if not attendance or attendance.user_id != user_id:
                return {
                    'success': False,
                    'message': 'Record absensi tidak ditemukan'
                }
            
            if not attendance.clock_out:
                return {
                    'success': False,
                    'message': 'Anda belum melakukan clock out'
                }
            
            # Cek apakah sudah ada overtime request untuk attendance ini
            existing_request = OvertimeRequest.query.filter(
                OvertimeRequest.attendance_id == attendance_id
            ).first()
            
            if existing_request:
                return {
                    'success': False,
                    'message': 'Overtime request sudah ada untuk hari ini'
                }
            
            # Buat overtime request
            overtime_request = OvertimeRequest(
                user_id=user_id,
                attendance_id=attendance_id,
                requested_hours=requested_hours,
                reason=reason,
                task_description=task_description,
                status='pending'
            )
            
            db.session.add(overtime_request)
            db.session.commit()
            
            logger.info(f"Overtime request submitted by user {user_id}")
            
            return {
                'success': True,
                'message': 'Pengajuan overtime berhasil dikirim',
                'data': overtime_request.to_dict()
            }
            
        except Exception as e:
            logger.error(f"Error submit overtime request: {str(e)}")
            db.session.rollback()
            return {
                'success': False,
                'message': f'Terjadi kesalahan: {str(e)}'
            }
    
    def approve_overtime_request(self, overtime_id: int, approver_id: int, 
                               action: str, actual_hours: float = None,
                               rejection_reason: str = None) -> Dict:
        """Approval overtime (approve/reject)"""
        try:
            overtime_request = OvertimeRequest.query.get(overtime_id)
            if not overtime_request:
                return {
                    'success': False,
                    'message': 'Pengajuan overtime tidak ditemukan'
                }
            
            if action == 'approve':
                overtime_request.approve(approver_id, actual_hours)
                message = 'Pengajuan overtime berhasil disetujui'
            elif action == 'reject':
                if not rejection_reason:
                    return {
                        'success': False,
                        'message': 'Alasan penolakan harus diisi'
                    }
                overtime_request.reject(approver_id, rejection_reason)
                message = 'Pengajuan overtime ditolak'
            else:
                return {
                    'success': False,
                    'message': 'Action tidak valid'
                }
            
            logger.info(f"Overtime request {overtime_id} {action}ed by {approver_id}")
            
            return {
                'success': True,
                'message': message,
                'data': overtime_request.to_dict()
            }
            
        except Exception as e:
            logger.error(f"Error approve overtime request: {str(e)}")
            db.session.rollback()
            return {
                'success': False,
                'message': f'Terjadi kesalahan: {str(e)}'
            }
    
    def get_dashboard_stats(self, user_id: int = None, start_date: date = None, 
                           end_date: date = None) -> Dict:
        """Stats untuk dashboard"""
        try:
            if not start_date:
                start_date = date.today().replace(day=1)  # Awal bulan
            if not end_date:
                end_date = date.today()  # Hari ini
            
            # Query base dengan error handling
            try:
                query = AttendanceRecord.query.filter(
                    db.func.date(AttendanceRecord.clock_in) >= start_date,
                    db.func.date(AttendanceRecord.clock_in) <= end_date
                )
                
                if user_id:
                    query = query.filter(AttendanceRecord.user_id == user_id)
                
                records = query.all()
            except Exception as e:
                logger.error(f"Error querying dashboard stats: {str(e)}", exc_info=True)
                records = []
            
            # Hitung stats dengan error handling
            try:
                total_days = len(records)
                present_days = len([r for r in records if r.status == 'present'])
                late_days = len([r for r in records if r.status == 'late'])
                absent_days = len([r for r in records if r.status == 'absent'])
                half_day_days = len([r for r in records if r.status == 'half_day'])
                
                # Hitung total jam kerja dengan safe access
                total_work_hours = sum([(r.work_duration_minutes or 0) for r in records]) / 60
                total_overtime_hours = sum([(r.overtime_minutes or 0) for r in records]) / 60
                
                # Hitung attendance rate
                working_days = (end_date - start_date).days + 1
                attendance_rate = round((present_days + late_days) / working_days * 100, 2) if working_days > 0 else 0
                
                return {
                    'period': {
                        'start_date': start_date.isoformat(),
                        'end_date': end_date.isoformat(),
                        'working_days': working_days
                    },
                    'attendance': {
                        'total_days': total_days,
                        'present_days': present_days,
                        'late_days': late_days,
                        'absent_days': absent_days,
                        'half_day_days': half_day_days,
                        'attendance_rate': attendance_rate
                    },
                    'work_hours': {
                        'total_work_hours': round(total_work_hours, 2),
                        'total_overtime_hours': round(total_overtime_hours, 2),
                        'average_daily_hours': round(total_work_hours / total_days, 2) if total_days > 0 else 0
                    }
                }
            except Exception as e:
                logger.error(f"Error calculating dashboard stats: {str(e)}", exc_info=True)
                # Return default values
                working_days = (end_date - start_date).days + 1
                return {
                    'period': {
                        'start_date': start_date.isoformat(),
                        'end_date': end_date.isoformat(),
                        'working_days': working_days
                    },
                    'attendance': {
                        'total_days': 0,
                        'present_days': 0,
                        'late_days': 0,
                        'absent_days': 0,
                        'half_day_days': 0,
                        'attendance_rate': 0
                    },
                    'work_hours': {
                        'total_work_hours': 0,
                        'total_overtime_hours': 0,
                        'average_daily_hours': 0
                    }
                }
            
        except (OperationalError, pymysql.err.OperationalError) as e:
            if is_encryption_error(e):
                logger.error(f"Error get dashboard stats - Database encryption error: {str(e)}", exc_info=True)
                raise ValueError("DATABASE_ENCRYPTION_ERROR")
            logger.error(f"Error get dashboard stats - Database error: {str(e)}", exc_info=True)
            return {}
        except Exception as e:
            logger.error(f"Error get dashboard stats: {str(e)}", exc_info=True)
            return {}
    
    def export_to_excel(self, records: List[Dict], filename: str = None) -> BytesIO:
        """Export attendance data ke Excel"""
        try:
            if not filename:
                filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance Report"
            
            # Header
            headers = [
                'No', 'Nama Karyawan', 'Tanggal', 'Clock In', 'Clock Out',
                'Durasi Kerja (Jam)', 'Overtime (Jam)', 'Status', 'Alamat Clock In', 'Alamat Clock Out', 'Catatan'
            ]
            
            # Style untuk header
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Set header
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Data
            for row, record in enumerate(records, 2):
                ws.cell(row=row, column=1, value=row-1)
                ws.cell(row=row, column=2, value=record.get('user_name', ''))
                ws.cell(row=row, column=3, value=record.get('clock_in', '').split('T')[0] if record.get('clock_in') else '')
                ws.cell(row=row, column=4, value=record.get('clock_in', '').split('T')[1][:8] if record.get('clock_in') else '')
                ws.cell(row=row, column=5, value=record.get('clock_out', '').split('T')[1][:8] if record.get('clock_out') else '')
                ws.cell(row=row, column=6, value=record.get('work_duration_hours', 0))
                ws.cell(row=row, column=7, value=record.get('overtime_hours', 0))
                ws.cell(row=row, column=8, value=record.get('status', ''))
                ws.cell(row=row, column=9, value=record.get('clock_in_address', ''))
                ws.cell(row=row, column=10, value=record.get('clock_out_address', ''))
                ws.cell(row=row, column=11, value=record.get('notes', ''))
            
            # Auto adjust column width
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output
            
        except Exception as e:
            logger.error(f"Error export to Excel: {str(e)}")
            raise e
    
    def export_to_pdf(self, records: List[Dict], filename: str = None) -> BytesIO:
        """Export attendance data ke PDF"""
        try:
            if not filename:
                filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            output = BytesIO()
            doc = SimpleDocTemplate(output, pagesize=A4)
            
            # Styles
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center
            )
            
            # Title
            title = Paragraph("Laporan Absensi Karyawan", title_style)
            
            # Table data
            table_data = [['No', 'Nama', 'Tanggal', 'Clock In', 'Clock Out', 'Durasi', 'Status']]
            
            for i, record in enumerate(records, 1):
                clock_in_date = record.get('clock_in', '').split('T')[0] if record.get('clock_in') else ''
                clock_in_time = record.get('clock_in', '').split('T')[1][:8] if record.get('clock_in') else ''
                clock_out_time = record.get('clock_out', '').split('T')[1][:8] if record.get('clock_out') else ''
                duration = f"{record.get('work_duration_hours', 0):.1f} jam"
                
                table_data.append([
                    str(i),
                    record.get('user_name', ''),
                    clock_in_date,
                    clock_in_time,
                    clock_out_time,
                    duration,
                    record.get('status', '')
                ])
            
            # Create table
            table = Table(table_data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            # Build PDF
            elements = [title, Spacer(1, 12), table]
            doc.build(elements)
            
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error export to PDF: {str(e)}")
            raise e

# Singleton instance
attendance_service = AttendanceService()
