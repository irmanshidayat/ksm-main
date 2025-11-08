#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Controller untuk Stok Barang - KSM Main Backend
Controller untuk mengelola API endpoint stok barang
"""

from flask import Blueprint, jsonify, request, send_file
from config.database import db
from services.stok_barang_service import StokBarangService
import logging
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# Setup logging
logger = logging.getLogger(__name__)

stok_barang_bp = Blueprint('stok_barang', __name__)

# Initialize service
stok_service = StokBarangService(db.session)

@stok_barang_bp.route('/health', methods=['GET'])
def health():
    """Health check endpoint"""
    return jsonify({"status": "ok", "service": "stok_barang"})

@stok_barang_bp.route('/dashboard', methods=['GET'])
def get_dashboard():
    """Endpoint untuk mengambil data dashboard stok barang"""
    try:
        logger.info("Mengambil data dashboard stok barang")
        data = stok_service.get_dashboard_data()
        return jsonify({
            "success": True,
            "data": data,
            "message": "Data dashboard berhasil diambil"
        })
    except Exception as e:
        logger.error(f"Error mengambil data dashboard: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Gagal mengambil data dashboard"
        }), 500

@stok_barang_bp.route('/barang', methods=['GET'])
def get_all_barang():
    """Endpoint untuk mengambil daftar barang dengan pagination, filter, dan search"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 10, type=int)
        search = request.args.get('search', type=str)
        kategori_id = request.args.get('kategori_id', type=int)
        stok_min = request.args.get('stok_min', type=int)
        stok_max = request.args.get('stok_max', type=int)
        harga_min = request.args.get('harga_min', type=float)
        harga_max = request.args.get('harga_max', type=float)
        sort_by = request.args.get('sort_by', 'nama_barang', type=str)
        sort_order = request.args.get('sort_order', 'asc', type=str)
        
        logger.info(f"Mengambil daftar barang - page: {page}, per_page: {per_page}, search: {search}")
        data = stok_service.get_all_barang(
            page=page, per_page=per_page, search=search, kategori_id=kategori_id,
            stok_min=stok_min, stok_max=stok_max, harga_min=harga_min, harga_max=harga_max,
            sort_by=sort_by, sort_order=sort_order
        )
        
        return jsonify({
            "success": True,
            "data": data,
            "message": "Daftar barang berhasil diambil"
        })
    except Exception as e:
        logger.error(f"Error mengambil daftar barang: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Gagal mengambil daftar barang"
        }), 500

@stok_barang_bp.route('/barang/<int:barang_id>', methods=['GET'])
def get_barang_by_id(barang_id):
    """Endpoint untuk mengambil detail barang berdasarkan ID"""
    try:
        logger.info(f"Mengambil detail barang ID: {barang_id}")
        data = stok_service.get_barang_by_id(barang_id)
        
        if not data:
            return jsonify({
                "success": False,
                "error": "Barang tidak ditemukan",
                "message": "Barang dengan ID tersebut tidak ditemukan"
            }), 404
        
        return jsonify({
            "success": True,
            "data": data,
            "message": "Detail barang berhasil diambil"
        })
    except Exception as e:
        logger.error(f"Error mengambil detail barang: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Gagal mengambil detail barang"
        }), 500

@stok_barang_bp.route('/barang/by-code/<string:kode_barang>', methods=['GET'])
def get_barang_by_code(kode_barang):
    """Endpoint untuk mengambil detail barang berdasarkan kode barang"""
    try:
        logger.info(f"Mengambil detail barang dengan kode: {kode_barang}")
        data = stok_service.get_barang_by_code(kode_barang)
        
        if not data:
            return jsonify({
                "success": False,
                "error": "Barang tidak ditemukan",
                "message": f"Barang dengan kode {kode_barang} tidak ditemukan"
            }), 404
        
        return jsonify({
            "success": True,
            "data": data,
            "message": "Detail barang berhasil diambil"
        })
    except Exception as e:
        logger.error(f"Error mengambil detail barang: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Gagal mengambil detail barang"
        }), 500

@stok_barang_bp.route('/barang', methods=['POST'])
def create_barang():
    """Endpoint untuk membuat barang baru"""
    try:
        data = request.get_json()
        
        # Validasi data yang diperlukan
        required_fields = ['kode_barang', 'nama_barang', 'satuan']
        for field in required_fields:
            if not data.get(field):
                return jsonify({
                    "success": False,
                    "error": f"Field {field} diperlukan",
                    "message": f"Field {field} tidak boleh kosong"
                }), 400
        
        logger.info(f"Membuat barang baru: {data['nama_barang']}")
        result = stok_service.create_barang(data)
        
        return jsonify({
            "success": True,
            "data": result,
            "message": "Barang berhasil dibuat"
        }), 201
    except Exception as e:
        logger.error(f"Error membuat barang: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Gagal membuat barang"
        }), 500

@stok_barang_bp.route('/barang/<int:barang_id>', methods=['PUT'])
def update_barang(barang_id):
    """Endpoint untuk update data barang"""
    try:
        data = request.get_json()
        
        logger.info(f"Update barang ID: {barang_id}")
        result = stok_service.update_barang(barang_id, data)
        
        return jsonify({
            "success": True,
            "data": result,
            "message": "Barang berhasil diupdate"
        })
    except Exception as e:
        logger.error(f"Error update barang: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Gagal update barang"
        }), 500

@stok_barang_bp.route('/barang/<int:barang_id>', methods=['DELETE'])
def delete_barang(barang_id):
    """Endpoint untuk menghapus barang (soft delete)"""
    try:
        logger.info(f"Menghapus barang ID: {barang_id}")
        result = stok_service.delete_barang(barang_id)
        
        return jsonify({
            "success": True,
            "data": {"deleted": result},
            "message": "Barang berhasil dihapus"
        })
    except Exception as e:
        logger.error(f"Error menghapus barang: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Gagal menghapus barang"
        }), 500

@stok_barang_bp.route('/kategori', methods=['GET'])
def get_kategori_list():
    """Endpoint untuk mengambil daftar kategori barang"""
    try:
        logger.info("Mengambil daftar kategori barang")
        data = stok_service.get_kategori_list()
        
        return jsonify({
            "success": True,
            "data": data,
            "message": "Daftar kategori berhasil diambil"
        })
    except Exception as e:
        logger.error(f"Error mengambil daftar kategori: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Gagal mengambil daftar kategori"
        }), 500

@stok_barang_bp.route('/merk', methods=['GET'])
def get_merk_list():
    """Endpoint untuk mengambil daftar merk yang ada"""
    try:
        logger.info("Mengambil daftar merk")
        data = stok_service.get_merk_list()
        
        # Ensure data is always a list
        if not isinstance(data, list):
            data = []
        
        return jsonify({
            "success": True,
            "data": data,
            "message": "Daftar merk berhasil diambil"
        })
    except Exception as e:
        logger.error(f"Error mengambil daftar merk: {str(e)}")
        # Return empty list instead of 500 error
        return jsonify({
            "success": True,
            "data": [],
            "message": "Daftar merk kosong"
        })

@stok_barang_bp.route('/supplier', methods=['GET'])
def get_supplier_list():
    """Endpoint untuk mengambil daftar supplier"""
    try:
        logger.info("Mengambil daftar supplier")
        data = stok_service.get_supplier_list()
        
        return jsonify({
            "success": True,
            "data": data,
            "message": "Daftar supplier berhasil diambil"
        })
    except Exception as e:
        logger.error(f"Error mengambil daftar supplier: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Gagal mengambil daftar supplier"
        }), 500

@stok_barang_bp.route('/barang/export', methods=['GET'])
def export_barang_excel():
    """Endpoint untuk export data barang ke Excel"""
    try:
        # Get filter parameters
        search = request.args.get('search', type=str)
        kategori_id = request.args.get('kategori_id', type=int)
        stok_min = request.args.get('stok_min', type=int)
        stok_max = request.args.get('stok_max', type=int)
        harga_min = request.args.get('harga_min', type=float)
        harga_max = request.args.get('harga_max', type=float)
        sort_by = request.args.get('sort_by', 'nama_barang', type=str)
        sort_order = request.args.get('sort_order', 'asc', type=str)
        
        logger.info("Export data barang ke Excel")
        
        # Get all data (no pagination for export)
        data = stok_service.get_all_barang(
            page=1, per_page=10000, search=search, kategori_id=kategori_id,
            stok_min=stok_min, stok_max=stok_max, harga_min=harga_min, harga_max=harga_max,
            sort_by=sort_by, sort_order=sort_order
        )
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Daftar Barang"
        
        # Define headers
        headers = [
            'No', 'Kode Barang', 'Nama Barang', 'Kategori', 'Satuan', 
            'Harga per Unit', 'Jumlah Stok', 'Stok Minimum', 'Stok Maksimum', 
            'Status Stok', 'Deskripsi'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row, item in enumerate(data['items'], 2):
            # Determine stock status
            stok_info = item.get('stok', {})
            jumlah_stok = stok_info.get('jumlah_stok', 0)
            stok_minimum = stok_info.get('stok_minimum', 0)
            
            if jumlah_stok == 0:
                status_stok = "HABIS"
            elif jumlah_stok <= stok_minimum:
                status_stok = "MENIPIS"
            else:
                status_stok = "AMAN"
            
            # Get category name
            kategori_info = item.get('kategori', {})
            kategori_nama = kategori_info.get('nama_kategori', '') if kategori_info else ''
            
            row_data = [
                row - 1,  # No
                item.get('kode_barang', ''),
                item.get('nama_barang', ''),
                kategori_nama,
                item.get('satuan', ''),
                item.get('harga_per_unit', 0),
                jumlah_stok,
                stok_minimum,
                stok_info.get('stok_maksimum', ''),
                status_stok,
                item.get('deskripsi', '')
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"daftar_barang_{timestamp}.xlsx"
        
        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        logger.error(f"Error export Excel: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Gagal export data ke Excel"
        }), 500

@stok_barang_bp.route('/barang-masuk', methods=['GET'])
def get_barang_masuk_list():
    """Endpoint untuk mengambil daftar barang masuk"""
    try:
        logger.info("Mengambil daftar barang masuk")
        data = stok_service.get_barang_masuk_list()
        
        return jsonify({
            "success": True,
            "data": data,
            "message": "Daftar barang masuk berhasil diambil"
        })
    except Exception as e:
        logger.error(f"Error mengambil daftar barang masuk: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Gagal mengambil daftar barang masuk"
        }), 500

@stok_barang_bp.route('/barang-masuk', methods=['POST'])
def create_barang_masuk():
    """Endpoint untuk membuat transaksi barang masuk baru"""
    try:
        data = request.get_json()
        
        # Validasi data yang diperlukan
        required_fields = ['barang_id', 'jumlah_masuk']
        for field in required_fields:
            if field not in data or data[field] is None:
                return jsonify({
                    "success": False,
                    "error": f"Field {field} diperlukan",
                    "message": f"Field {field} tidak boleh kosong"
                }), 400
        
        if data['jumlah_masuk'] <= 0:
            return jsonify({
                "success": False,
                "error": "Jumlah masuk harus lebih dari 0",
                "message": "Jumlah masuk harus lebih dari 0"
            }), 400
        
        logger.info(f"Membuat transaksi barang masuk baru: barang_id={data['barang_id']}, jumlah={data['jumlah_masuk']}")
        result = stok_service.create_barang_masuk(data)
        
        return jsonify({
            "success": True,
            "data": result,
            "message": "Barang masuk berhasil ditambahkan"
        }), 201
    except Exception as e:
        logger.error(f"Error membuat barang masuk: {str(e)}")
        return jsonify({
            "success": False,
            "error": str(e),
            "message": "Gagal membuat barang masuk"
        }), 500


