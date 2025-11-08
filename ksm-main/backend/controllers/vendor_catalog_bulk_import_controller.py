#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Vendor Catalog Bulk Import Controller - API Endpoints untuk bulk import vendor catalog items
Controller untuk mengelola bulk import barang vendor dari file Excel/CSV
"""

from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity, verify_jwt_in_request
from typing import List, Dict, Any
from datetime import datetime
import logging
import pandas as pd
import io
from functools import wraps

from config.database import db
from services.vendor_catalog_service import VendorCatalogService
from utils.serialization import serialize_models
from utils.security_validator import security_validator
from utils.rate_limiter import rate_limiter

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed, Excel template generation will not work")

# Buat Blueprint untuk Flask
vendor_catalog_bulk_import_bp = Blueprint('vendor_catalog_bulk_import', __name__)

# Custom decorator untuk mengizinkan OPTIONS request
def jwt_required_allow_options(f):
    """Decorator yang mengizinkan OPTIONS request untuk CORS preflight"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if request.method == 'OPTIONS':
            return '', 200
        try:
            verify_jwt_in_request()
            return f(*args, **kwargs)
        except Exception as e:
            return jsonify({'error': 'Authentication failed'}), 401
    return decorated_function

# ===== BULK IMPORT VALIDATION =====

@vendor_catalog_bulk_import_bp.route("/api/vendor-catalog/bulk-validate", methods=["POST"])
@jwt_required_allow_options
def validate_bulk_import():
    """Validate file untuk bulk import vendor catalog items"""
    try:
        # Check if file is provided
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'message': 'File tidak ditemukan'
            }), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({
                'success': False,
                'message': 'File tidak dipilih'
            }), 400
        
        # Read file based on extension
        file_extension = file.filename.split('.')[-1].lower()
        
        try:
            if file_extension in ['xlsx', 'xls']:
                # Reset file pointer
                file.seek(0)
                if file_extension == 'xlsx':
                    df = pd.read_excel(file, engine='openpyxl')
                else:
                    df = pd.read_excel(file, engine='xlrd')
            elif file_extension == 'csv':
                file.seek(0)
                df = pd.read_csv(file)
            else:
                return jsonify({
                    'success': False,
                    'message': 'Format file tidak didukung. Gunakan Excel (.xlsx, .xls) atau CSV (.csv)'
                }), 400
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Gagal membaca file: {str(e)}'
            }), 400
        
        # Validate required columns - sesuai dengan field yang diperlukan
        required_columns = ['Nama Barang', 'Spesifikasi Teknis', 'Vendor', 'Email', 'Quantity', 'Harga Satuan', 'Harga Total', 'Kategori', 'Merek', 'Status', 'Tanggal']
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return jsonify({
                'success': False,
                'message': f'Kolom yang diperlukan tidak ditemukan: {", ".join(missing_columns)}'
            }), 400
        
        # Validate data
        errors = []
        warnings = []
        preview_data = []
        
        for index, row in df.iterrows():
            row_errors = []
            
            # Validate required fields
            if pd.isna(row['Nama Barang']) or str(row['Nama Barang']).strip() == '':
                row_errors.append('Nama Barang tidak boleh kosong')
            
            if pd.isna(row['Vendor']) or str(row['Vendor']).strip() == '':
                row_errors.append('Vendor tidak boleh kosong')
            
            if pd.isna(row['Quantity']) or not str(row['Quantity']).isdigit():
                row_errors.append('Quantity harus berupa angka')
            
            if pd.isna(row['Harga Satuan']) or not str(row['Harga Satuan']).replace('.', '').replace(',', '').isdigit():
                row_errors.append('Harga Satuan harus berupa angka')
            
            if pd.isna(row['Harga Total']) or not str(row['Harga Total']).replace('.', '').replace(',', '').isdigit():
                row_errors.append('Harga Total harus berupa angka')
            
            # Validate date format (DD/MM/YYYY)
            if not pd.isna(row['Tanggal']):
                try:
                    from datetime import datetime
                    datetime.strptime(str(row['Tanggal']), '%d/%m/%Y')
                except ValueError:
                    row_errors.append('Tanggal harus dalam format DD/MM/YYYY')
            
            # TODO: Check if vendor exists in database by name
            
            if row_errors:
                errors.extend([f'Baris {index + 2}: {error}' for error in row_errors])
            else:
                # Add to preview data
                preview_data.append({
                    'nama_barang': str(row['Nama Barang']).strip(),
                    'vendor_name': str(row['Vendor']).strip(),
                    'quantity': int(row['Quantity']),
                    'harga_satuan': float(str(row['Harga Satuan']).replace(',', '')),
                    'harga_total': float(str(row['Harga Total']).replace(',', '')),
                    'kategori': str(row['Kategori']).strip() if not pd.isna(row['Kategori']) else '',
                    'merek': str(row['Merek']).strip() if not pd.isna(row['Merek']) else '',
                    'spesifikasi': str(row['Spesifikasi Teknis']).strip() if not pd.isna(row['Spesifikasi Teknis']) else '',
                    'status': str(row['Status']).strip() if not pd.isna(row['Status']) else 'submitted',
                    'tanggal': str(row['Tanggal']).strip() if not pd.isna(row['Tanggal']) else ''
                })
        
        return jsonify({
            'success': True,
            'data': {
                'isValid': len(errors) == 0,
                'errors': errors,
                'warnings': warnings,
                'preview': preview_data
            }
        })
        
    except Exception as e:
        logger.error(f"Error validating bulk import: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Terjadi kesalahan saat memvalidasi file: {str(e)}'
        }), 500

# ===== BULK IMPORT EXECUTION =====

@vendor_catalog_bulk_import_bp.route("/api/vendor-catalog/bulk-import", methods=["POST"])
@jwt_required_allow_options
def execute_bulk_import():
    """Execute bulk import vendor catalog items"""
    try:
        # Check if file is provided
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'message': 'File tidak ditemukan'
            }), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({
                'success': False,
                'message': 'File tidak dipilih'
            }), 400
        
        # Read file
        file_extension = file.filename.split('.')[-1].lower()
        
        try:
            if file_extension in ['xlsx', 'xls']:
                # Reset file pointer
                file.seek(0)
                if file_extension == 'xlsx':
                    df = pd.read_excel(file, engine='openpyxl')
                else:
                    df = pd.read_excel(file, engine='xlrd')
            elif file_extension == 'csv':
                file.seek(0)
                df = pd.read_csv(file)
            else:
                return jsonify({
                    'success': False,
                    'message': 'Format file tidak didukung'
                }), 400
        except Exception as e:
            return jsonify({
                'success': False,
                'message': f'Gagal membaca file: {str(e)}'
            }), 400
        
        # Process data using service
        items_data = []
        
        for index, row in df.iterrows():
            # Validate row data
            if pd.isna(row['Nama Barang']) or str(row['Nama Barang']).strip() == '':
                continue  # Skip invalid rows
            
            item_data = {
                'nama_barang': str(row['Nama Barang']).strip(),
                'vendor_name': str(row['Vendor']).strip(),
                'email': str(row['Email']).strip() if not pd.isna(row['Email']) else '',
                'quantity': int(row['Quantity']),
                'harga_satuan': float(str(row['Harga Satuan']).replace(',', '')),
                'harga_total': float(str(row['Harga Total']).replace(',', '')),
                'kategori': str(row['Kategori']).strip() if not pd.isna(row['Kategori']) else '',
                'merek': str(row['Merek']).strip() if not pd.isna(row['Merek']) else '',
                'spesifikasi': str(row['Spesifikasi Teknis']).strip() if not pd.isna(row['Spesifikasi Teknis']) else '',
                'status': str(row['Status']).strip() if not pd.isna(row['Status']) else 'submitted',
                'tanggal': str(row['Tanggal']).strip() if not pd.isna(row['Tanggal']) else ''
            }
            items_data.append(item_data)
        
        # Use service to bulk import
        service = VendorCatalogService(db.session)
        result = service.bulk_import_vendor_catalog_items(items_data)
        
        if result['success']:
            successful_imports = result['successful_imports']
            failed_imports = result['failed_imports']
            errors = result['errors']
            imported_items = result['imported_items']
        else:
            return jsonify({
                'success': False,
                'message': result['message']
            }), 500
        
        return jsonify({
            'success': True,
            'data': {
                'total_rows': len(df),
                'successful_imports': successful_imports,
                'failed_imports': failed_imports,
                'errors': errors,
                'imported_items': imported_items
            }
        })
        
    except Exception as e:
        logger.error(f"Error executing bulk import: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Terjadi kesalahan saat mengimport data: {str(e)}'
        }), 500

# ===== SINGLE ITEM IMPORT =====

@vendor_catalog_bulk_import_bp.route("/api/vendor-catalog/bulk-import-single", methods=["POST"])
@jwt_required_allow_options
def import_single_vendor_catalog_item():
    """Import single vendor catalog item (for form submission)"""
    try:
        item_data = request.get_json()
        if not item_data:
            return jsonify({
                'success': False,
                'message': 'Item data tidak ditemukan'
            }), 400
        
        # Validate required fields
        required_fields = ['nama_barang', 'vendor_name', 'quantity', 'harga_satuan', 'harga_total']
        for field in required_fields:
            if field not in item_data or not item_data[field]:
                return jsonify({
                    'success': False,
                    'message': f'Field {field} is required'
                }), 400
        
        # Use service to import single item
        service = VendorCatalogService(db.session)
        result = service.bulk_import_vendor_catalog_items([item_data])
        
        if result['success']:
            return jsonify({
                'success': True,
                'data': {
                    'imported_items': result['imported_items'],
                    'successful_imports': result['successful_imports']
                }
            })
        else:
            return jsonify({
                'success': False,
                'message': result.get('message', 'Gagal mengimport item')
            }), 400
        
    except Exception as e:
        logger.error(f"❌ Error importing single item: {str(e)}")
        return jsonify({
            'success': False,
            'message': str(e)
        }), 500

# ===== TEMPLATE DOWNLOAD =====

@vendor_catalog_bulk_import_bp.route("/api/vendor-catalog/bulk-import/template", methods=["GET"])
@jwt_required_allow_options
def download_template():
    """Download template Excel untuk bulk import vendor catalog items"""
    try:
        if not OPENPYXL_AVAILABLE:
            return jsonify({
                'success': False,
                'message': 'openpyxl tidak terinstall, tidak dapat generate template Excel'
            }), 500

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Template Import"

        # Define headers sesuai dengan required_columns
        headers = [
            'Nama Barang',
            'Spesifikasi Teknis',
            'Vendor',
            'Email',
            'Quantity',
            'Harga Satuan',
            'Harga Total',
            'Kategori',
            'Merek',
            'Status',
            'Tanggal'
        ]

        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Add example row (optional - bisa dikosongkan atau diisi contoh)
        example_row = [
            'Printer HP LaserJet',
            'Laser printer, 1200x1200 dpi, USB/Network',
            'PT Vendor ABC',
            'vendor@example.com',
            2,
            1500000,
            3000000,
            'Elektronik',
            'HP',
            'submitted',
            datetime.now().strftime('%Y-%m-%d')
        ]

        for col, value in enumerate(example_row, 1):
            ws.cell(row=2, column=col, value=value)

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d")
        filename = f"template_import_barang_vendor_{timestamp}.xlsx"

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        logger.error(f"❌ Error generating template: {str(e)}")
        return jsonify({
            'success': False,
            'message': f'Gagal membuat template: {str(e)}'
        }), 500

# ===== HEALTH CHECK =====

@vendor_catalog_bulk_import_bp.route("/api/vendor-catalog/bulk-import/health", methods=["GET"])
def health_check():
    """Health check untuk bulk import service"""
    return jsonify({
        'status': 'healthy',
        'service': 'vendor_catalog_bulk_import',
        'timestamp': datetime.utcnow().isoformat()
    })
